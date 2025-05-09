import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.utils import timezone

def generate_sales_report_excel(start_date, end_date, sales, summary_data):
    """Generate an Excel report for sales."""
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create styles
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    normal_font = Font(name='Arial', size=11)
    date_font = Font(name='Arial', size=11)
    money_font = Font(name='Arial', size=11)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === Summary Sheet ===
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Title
    summary_sheet['A1'] = "Fruit Juice Production Ltd"
    summary_sheet['A1'].font = Font(name='Arial', bold=True, size=16)
    summary_sheet.merge_cells('A1:E1')
    summary_sheet['A1'].alignment = Alignment(horizontal='center')
    
    summary_sheet['A2'] = "Sales Report"
    summary_sheet['A2'].font = Font(name='Arial', bold=True, size=14)
    summary_sheet.merge_cells('A2:E2')
    summary_sheet['A2'].alignment = Alignment(horizontal='center')
    
    summary_sheet['A3'] = f"Period: {start_date.strftime('%d %b, %Y')} to {end_date.strftime('%d %b, %Y')}"
    summary_sheet['A3'].font = Font(name='Arial', size=12)
    summary_sheet.merge_cells('A3:E3')
    summary_sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Summary data
    summary_sheet['A5'] = "Summary Statistics"
    summary_sheet['A5'].font = Font(name='Arial', bold=True, size=12)
    summary_sheet.merge_cells('A5:B5')
    
    summary_sheet['A6'] = "Total Sales:"
    summary_sheet['A6'].alignment = Alignment(horizontal='right')
    summary_sheet['B6'] = summary_data['total_sales']
    summary_sheet['B6'].number_format = 'GHS#,##0.00'
    
    summary_sheet['A7'] = "Number of Sales:"
    summary_sheet['A7'].alignment = Alignment(horizontal='right')
    summary_sheet['B7'] = summary_data['total_count']
    
    summary_sheet['A8'] = "Average Sale Value:"
    summary_sheet['A8'].alignment = Alignment(horizontal='right')
    summary_sheet['B8'] = summary_data['avg_sale']
    summary_sheet['B8'].number_format = 'GHS#,##0.00'
    
    summary_sheet['A9'] = "Total Products Sold:"
    summary_sheet['A9'].alignment = Alignment(horizontal='right')
    summary_sheet['B9'] = summary_data['total_products_sold']
    
    # Format and size summary section
    for row in range(6, 10):
        for col in range(1, 3):
            summary_sheet.cell(row=row, column=col).border = thin_border
            summary_sheet.cell(row=row, column=col).font = normal_font
    
    # Status breakdown
    if 'sales_by_status' in summary_data:
        summary_sheet['A11'] = "Sales by Status"
        summary_sheet['A11'].font = Font(name='Arial', bold=True, size=12)
        summary_sheet.merge_cells('A11:C11')
        
        # Header
        summary_sheet['A12'] = "Status"
        summary_sheet['B12'] = "Count"
        summary_sheet['C12'] = "Total Amount"
        
        for col in range(1, 4):
            cell = summary_sheet.cell(row=12, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        row = 13
        for status in summary_data['sales_by_status']:
            summary_sheet.cell(row=row, column=1).value = status['status'].title()
            summary_sheet.cell(row=row, column=2).value = status['count']
            summary_sheet.cell(row=row, column=3).value = status['total']
            summary_sheet.cell(row=row, column=3).number_format = 'GHS#,##0.00'
            
            for col in range(1, 4):
                summary_sheet.cell(row=row, column=col).border = thin_border
                summary_sheet.cell(row=row, column=col).font = normal_font
            
            row += 1
    
    # Top products
    if 'top_products' in summary_data:
        # Leave some space after the status table
        start_row = row + 2
        
        summary_sheet.cell(row=start_row, column=1).value = "Top 10 Products"
        summary_sheet.cell(row=start_row, column=1).font = Font(name='Arial', bold=True, size=12)
        summary_sheet.merge_cells(f'A{start_row}:D{start_row}')
        
        # Header row
        summary_sheet.cell(row=start_row+1, column=1).value = "Product"
        summary_sheet.cell(row=start_row+1, column=2).value = "Quantity"
        summary_sheet.cell(row=start_row+1, column=3).value = "Revenue"
        summary_sheet.cell(row=start_row+1, column=4).value = "Avg. Price"
        
        for col in range(1, 5):
            cell = summary_sheet.cell(row=start_row+1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        product_row = start_row + 2
        for product in summary_data['top_products'][:10]:  # Limit to top 10
            summary_sheet.cell(row=product_row, column=1).value = product['product__name']
            summary_sheet.cell(row=product_row, column=2).value = product['quantity']
            summary_sheet.cell(row=product_row, column=3).value = product['revenue']
            summary_sheet.cell(row=product_row, column=3).number_format = 'GHS#,##0.00'
            summary_sheet.cell(row=product_row, column=4).value = product['avg_price']
            summary_sheet.cell(row=product_row, column=4).number_format = 'GHS#,##0.00'
            
            for col in range(1, 5):
                summary_sheet.cell(row=product_row, column=col).border = thin_border
                summary_sheet.cell(row=product_row, column=col).font = normal_font
            
            product_row += 1
    
    # Adjust column widths
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        summary_sheet.column_dimensions[column_letter].width = 20
    
    # === Sales Details Sheet ===
    sales_sheet = wb.create_sheet(title="Sales List")
    
    # Header
    sales_sheet['A1'] = "Invoice #"
    sales_sheet['B1'] = "Date"
    sales_sheet['C1'] = "Customer"
    sales_sheet['D1'] = "Amount"
    sales_sheet['E1'] = "Status"
    sales_sheet['F1'] = "Payment Status"
    
    for col in range(1, 7):
        cell = sales_sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for idx, sale in enumerate(sales, start=2):
        sales_sheet.cell(row=idx, column=1).value = sale.invoice_number
        sales_sheet.cell(row=idx, column=2).value = sale.sale_date
        sales_sheet.cell(row=idx, column=2).number_format = 'YYYY-MM-DD'
        sales_sheet.cell(row=idx, column=3).value = sale.customer.name
        sales_sheet.cell(row=idx, column=4).value = sale.total_amount
        sales_sheet.cell(row=idx, column=4).number_format = 'GHS#,##0.00'
        sales_sheet.cell(row=idx, column=5).value = sale.status.title()
        sales_sheet.cell(row=idx, column=6).value = sale.payment_status.title()
        
        for col in range(1, 7):
            sales_sheet.cell(row=idx, column=col).border = thin_border
            sales_sheet.cell(row=idx, column=col).font = normal_font
    
    # Adjust column widths
    sales_sheet.column_dimensions['A'].width = 15
    sales_sheet.column_dimensions['B'].width = 12
    sales_sheet.column_dimensions['C'].width = 30
    sales_sheet.column_dimensions['D'].width = 15
    sales_sheet.column_dimensions['E'].width = 15
    sales_sheet.column_dimensions['F'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output